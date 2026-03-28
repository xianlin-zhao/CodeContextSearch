import argparse
import csv
import os
import time
import traceback
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterator, Set, Tuple

from main_mopidy import main as summarize_project


SKIP_PROJECT_NAMES = {
    "Telethon",
    "pyramid",
    "pyinfra",
    "kinto",
    "capirca",
    "sqlitedict",
}

SKIP_PROJECT2_NAMES = {
    "Jinja2",
    "psd-tools",
    "bentoml",
    "datasets",
    "diffusers",
    "pyOpenSSL",
    "pandas-profiling",
    "ydata-profiling",
    "exodus-bundler",
    "viztracer",
    "mistune",
}


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def discover_projects_from_excel(projects_excel: Path) -> Iterator[Tuple[str, str, Path]]:
    """Yield (category, project_name, project_path) from an Excel file."""
    log(f"Loading projects Excel: {projects_excel}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read .xlsx files. Please install it with: pip install openpyxl") from exc

    workbook = load_workbook(filename=str(projects_excel), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            log("Excel appears empty (no header row).")
            return

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]

        try:
            project_name_idx = headers.index("project_name")
            project_path_idx = headers.index("project_path")
        except ValueError as exc:
            raise ValueError(
                f"Excel file {projects_excel} must contain 'project_name' and 'project_path' columns in the header row"
            ) from exc

        scanned_rows = 0
        yielded_projects = 0
        log("Excel header parsed, scanning project rows...")
        for row in rows:
            scanned_rows += 1
            if scanned_rows % 1000 == 0:
                log(f"Excel scan progress: rows={scanned_rows}, valid_projects={yielded_projects}")

            if row is None:
                continue

            project_name = str(row[project_name_idx]).strip() if project_name_idx < len(row) and row[project_name_idx] is not None else ""
            project_path_raw = str(row[project_path_idx]).strip() if project_path_idx < len(row) and row[project_path_idx] is not None else ""

            if not project_name or not project_path_raw:
                continue

            project_path = Path(project_path_raw).expanduser()
            if not project_path.is_absolute():
                project_path = (projects_excel.parent / project_path).resolve()
            else:
                project_path = project_path.resolve()

            if len(project_path.parents) >= 2 and project_path.parents[1].name:
                category = project_path.parents[1].name
            else:
                category = project_path.parent.name
            yielded_projects += 1
            yield category, project_name, project_path
        log(f"Excel scan complete: rows={scanned_rows}, valid_projects={yielded_projects}")
    finally:
        workbook.close()


def run_one_project(project_path: Path, output_dir: Path) -> Dict[str, object]:
    """Run summarization for one project and normalize result schema."""
    os.makedirs(output_dir, exist_ok=True)
    result = summarize_project(project_root=str(project_path), output_dir=str(output_dir))

    if result is None:
        return {
            "status": "failed",
            "reason": "main_mopidy returned None",
            "total_functions": 0,
            "total_features": 0,
        }

    return {
        "status": result.get("status", "unknown"),
        "reason": result.get("reason", ""),
        "total_functions": int(result.get("total_functions", 0) or 0),
        "total_features": int(result.get("total_features", 0) or 0),
        "description_issue_features": int(result.get("description_issue_features", 0) or 0),
        "description_fallback_features": int(result.get("description_fallback_features", 0) or 0),
        "description_issue_feature_ids": str(result.get("description_issue_feature_ids", [])),
    }


def load_processed_projects(report_file: Path) -> Set[Tuple[str, str]]:
    processed: Set[Tuple[str, str]] = set()
    if not report_file.exists():
        return processed

    with open(report_file, "r", encoding="utf-8", newline="") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            category = (row.get("category") or "").strip()
            project = (row.get("project") or "").strip()
            if category and project:
                processed.add((category, project))
    return processed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch-run main_mopidy for DevEval projects.")
    parser.add_argument(
        "--projects-excel",
        default="/data/data_public/riverbag/CodeContextSearch/docs/deveval_project_path_stats.xlsx",
        help="Excel file containing at least project_name and project_path columns.",
    )
    parser.add_argument(
        "--output-root",
        default="/data/data_public/riverbag/testRepoSummaryOut/DevEval",
        help="Root output directory for project summaries.",
    )
    parser.add_argument(
        "--report-file",
        default="/data/data_public/riverbag/testRepoSummaryOut/DevEval/run_summary.csv",
        help="CSV file to store per-project function/feature statistics.",
    )
    parser.add_argument(
        "--max-new-runs",
        type=int,
        default=0,
        help="Run at most N newly executed projects (skipped projects do not count). 0 means no limit.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=1,
        help="Number of parallel workers. 1 means serial execution.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    log("Batch run started.")
    log(
        "Arguments: "
        f"projects_excel={args.projects_excel}, output_root={args.output_root}, "
        f"report_file={args.report_file}, max_new_runs={args.max_new_runs}, workers={args.workers}"
    )

    if args.workers < 1:
        raise ValueError(f"--workers must be >= 1, got: {args.workers}")

    projects_excel = Path(args.projects_excel).resolve()
    output_root = Path(args.output_root).resolve()
    report_file = Path(args.report_file).resolve()

    if not projects_excel.exists() or not projects_excel.is_file():
        raise FileNotFoundError(f"Invalid projects Excel file: {projects_excel}")

    os.makedirs(output_root, exist_ok=True)
    os.makedirs(report_file.parent, exist_ok=True)

    log("Reading projects from Excel into memory...")
    projects = list(discover_projects_from_excel(projects_excel))
    log(f"Loaded {len(projects)} projects from Excel.")

    processed_projects = load_processed_projects(report_file)
    log(f"Loaded {len(processed_projects)} processed project records from report.")

    effective_new_runs_limit = args.max_new_runs if args.max_new_runs and args.max_new_runs > 0 else 0

    fieldnames = [
        "category",
        "project",
        "project_root",
        "output_dir",
        "status",
        "reason",
        "total_functions",
        "total_features",
        "description_issue_features",
        "description_fallback_features",
        "description_issue_feature_ids",
        "start_time",
        "end_time",
        "duration_seconds",
    ]

    write_header = not report_file.exists() or report_file.stat().st_size == 0
    with open(report_file, "a", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        if write_header:
            writer.writeheader()

        total = len(projects)
        pending_projects = []
        log("Filtering projects to build pending run list...")
        for idx, (category, project_name, project_path) in enumerate(projects, start=1):
            if idx % 1000 == 0:
                log(f"Filter progress: inspected={idx}/{total}, pending={len(pending_projects)}")

            if effective_new_runs_limit > 0 and len(pending_projects) >= effective_new_runs_limit:
                print(f"Reached run limit={effective_new_runs_limit} (new runs only). Stop batch early.")
                break

            project_output_dir = output_root / category / project_name

            if project_name in SKIP_PROJECT_NAMES or project_name in SKIP_PROJECT2_NAMES:
                print(f"[{idx}/{total}] Skip {category}/{project_name} | in skip lists by project name")
                continue

            if (category, project_name) in processed_projects:
                print(f"[{idx}/{total}] Skip {category}/{project_name} | already in report")
                continue

            if project_output_dir.exists():
                print(f"[{idx}/{total}] Skip {category}/{project_name} | output exists: {project_output_dir}")
                continue

            pending_projects.append(
                {
                    "idx": idx,
                    "category": category,
                    "project_name": project_name,
                    "project_path": project_path,
                    "project_output_dir": project_output_dir,
                }
            )

        log(f"Pending projects ready: {len(pending_projects)}")

        if args.workers == 1:
            log("Execution mode: serial (workers=1)")
            for item in pending_projects:
                idx = int(item["idx"])
                category = str(item["category"])
                project_name = str(item["project_name"])
                project_path = Path(item["project_path"])
                project_output_dir = Path(item["project_output_dir"])

                print(f"[{idx}/{total}] Running {category}/{project_name}")
                start_ts = datetime.now()
                start_time = time.time()
                status = "failed"
                reason = ""
                total_functions = 0
                total_features = 0
                description_issue_features = 0
                description_fallback_features = 0
                description_issue_feature_ids = "[]"

                try:
                    summary = run_one_project(project_path, project_output_dir)
                    status = summary["status"]
                    reason = str(summary["reason"])
                    total_functions = int(summary["total_functions"])
                    total_features = int(summary["total_features"])
                    description_issue_features = int(summary["description_issue_features"])
                    description_fallback_features = int(summary["description_fallback_features"])
                    description_issue_feature_ids = str(summary["description_issue_feature_ids"])
                except Exception as exc:
                    reason = f"{type(exc).__name__}: {exc}"
                    traceback.print_exc()

                end_ts = datetime.now()
                duration_seconds = round(time.time() - start_time, 3)

                row = {
                    "category": category,
                    "project": project_name,
                    "project_root": str(project_path),
                    "output_dir": str(project_output_dir),
                    "status": status,
                    "reason": reason,
                    "total_functions": total_functions,
                    "total_features": total_features,
                    "description_issue_features": description_issue_features,
                    "description_fallback_features": description_fallback_features,
                    "description_issue_feature_ids": description_issue_feature_ids,
                    "start_time": start_ts.isoformat(timespec="seconds"),
                    "end_time": end_ts.isoformat(timespec="seconds"),
                    "duration_seconds": duration_seconds,
                }
                writer.writerow(row)
                fp.flush()
                processed_projects.add((category, project_name))

                print(
                    f"[{idx}/{total}] Done {category}/{project_name} | "
                    f"status={status} functions={total_functions} features={total_features} "
                    f"desc_issues={description_issue_features}"
                )
        elif pending_projects:
            log(f"Execution mode: parallel (workers={args.workers})")
            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures: Dict[Future, Dict[str, object]] = {}
                for item in pending_projects:
                    idx = int(item["idx"])
                    category = str(item["category"])
                    project_name = str(item["project_name"])
                    project_path = Path(item["project_path"])
                    project_output_dir = Path(item["project_output_dir"])

                    print(f"[{idx}/{total}] Running {category}/{project_name}")
                    start_ts = datetime.now()
                    start_time = time.time()
                    future = executor.submit(run_one_project, project_path, project_output_dir)
                    futures[future] = {
                        "idx": idx,
                        "total": total,
                        "category": category,
                        "project_name": project_name,
                        "project_path": project_path,
                        "project_output_dir": project_output_dir,
                        "start_ts": start_ts,
                        "start_time": start_time,
                    }

                for future in as_completed(futures):
                    meta = futures[future]
                    idx = int(meta["idx"])
                    total_count = int(meta["total"])
                    category = str(meta["category"])
                    project_name = str(meta["project_name"])
                    project_path = Path(meta["project_path"])
                    project_output_dir = Path(meta["project_output_dir"])
                    start_ts = meta["start_ts"]
                    start_time = float(meta["start_time"])

                    status = "failed"
                    reason = ""
                    total_functions = 0
                    total_features = 0
                    description_issue_features = 0
                    description_fallback_features = 0
                    description_issue_feature_ids = "[]"

                    try:
                        summary = future.result()
                        status = summary["status"]
                        reason = str(summary["reason"])
                        total_functions = int(summary["total_functions"])
                        total_features = int(summary["total_features"])
                        description_issue_features = int(summary["description_issue_features"])
                        description_fallback_features = int(summary["description_fallback_features"])
                        description_issue_feature_ids = str(summary["description_issue_feature_ids"])
                    except Exception as exc:
                        reason = f"{type(exc).__name__}: {exc}"
                        traceback.print_exc()

                    end_ts = datetime.now()
                    duration_seconds = round(time.time() - start_time, 3)

                    row = {
                        "category": category,
                        "project": project_name,
                        "project_root": str(project_path),
                        "output_dir": str(project_output_dir),
                        "status": status,
                        "reason": reason,
                        "total_functions": total_functions,
                        "total_features": total_features,
                        "description_issue_features": description_issue_features,
                        "description_fallback_features": description_fallback_features,
                        "description_issue_feature_ids": description_issue_feature_ids,
                        "start_time": start_ts.isoformat(timespec="seconds"),
                        "end_time": end_ts.isoformat(timespec="seconds"),
                        "duration_seconds": duration_seconds,
                    }
                    writer.writerow(row)
                    fp.flush()
                    processed_projects.add((category, project_name))

                    print(
                        f"[{idx}/{total_count}] Done {category}/{project_name} | "
                        f"status={status} functions={total_functions} features={total_features} "
                        f"desc_issues={description_issue_features}"
                    )
        else:
            log("No pending projects to run after filtering.")

    log(f"Batch finished. Report saved to: {report_file}")


if __name__ == "__main__":
    main()
